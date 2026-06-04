import streamlit as st
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import matplotlib.patches as mpatches
import io
import time
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

# ── NEW: Excel Export ──
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference, LineChart
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── NEW: 3D ──
try:
    from mpl_toolkits.mplot3d import Axes3D
    from mpl_toolkits.mplot3d.art3d import Poly3DCollection
    THREED_OK = True
except ImportError:
    THREED_OK = False

st.set_page_config(page_title="Beam Simulator Pro", page_icon="🏗️", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Orbitron:wght@700;900&family=Rajdhani:wght@400;600&display=swap');

    /* ── Islands Dark / Antigravity Theme ── */
    .stApp {
        background: #0e1117;
        color: #e0e0e0;
        font-family: 'Rajdhani', sans-serif;
        position: relative;
        overflow-x: hidden;
    }

    /* Make sure all Streamlit content is above canvas */
    .main, .main .block-container,
    [data-testid="stAppViewContainer"] > section,
    [data-testid="block-container"],
    .stSidebar, header {
        position: relative !important;
        z-index: 10 !important;
    }

    /* Deep dark gradient base */
    .stApp::before {
        content: '';
        position: fixed; top: 0; left: 0; width: 100%; height: 100%;
        background:
            radial-gradient(ellipse at 0% 0%,   #1a1040 0%, transparent 50%),
            radial-gradient(ellipse at 100% 0%,  #0a1a2e 0%, transparent 50%),
            radial-gradient(ellipse at 100% 100%,#1a0a2e 0%, transparent 50%),
            radial-gradient(ellipse at 0% 100%,  #0a1a1a 0%, transparent 50%),
            radial-gradient(ellipse at 50% 50%,  #12082a 0%, transparent 70%);
        pointer-events: none; z-index: -2;
    }

    /* Floating glass panels effect */
    .stApp::after {
        content: '';
        position: fixed; top: 0; left: 0; width: 100%; height: 100%;
        background:
            radial-gradient(ellipse 40% 25% at 20% 35%, rgba(99,102,241,0.08) 0%, transparent 100%),
            radial-gradient(ellipse 35% 20% at 80% 65%, rgba(139,92,246,0.07) 0%, transparent 100%),
            radial-gradient(ellipse 50% 30% at 60% 20%, rgba(59,130,246,0.06) 0%, transparent 100%),
            radial-gradient(ellipse 30% 35% at 35% 80%, rgba(16,185,129,0.05) 0%, transparent 100%);
        pointer-events: none; z-index: -2;
        animation: islandsFloat 12s ease-in-out infinite alternate;
    }
    @keyframes islandsFloat {
        0%   { opacity: 0.7; transform: translateY(0px); }
        50%  { opacity: 1.0; transform: translateY(-6px); }
        100% { opacity: 0.8; transform: translateY(3px); }
    }

    /* ── Sidebar — glass panel ── */
    .stSidebar {
        background: rgba(15, 12, 30, 0.85) !important;
        border-right: 1px solid rgba(99,102,241,0.2) !important;
        -webkit-backdrop-filter: blur(20px) !important;
        backdrop-filter: blur(20px) !important;
        box-shadow: 4px 0 30px rgba(99,102,241,0.1) !important;
    }
    .stSidebar .stMarkdown h2 {
        background: linear-gradient(90deg, #00bfff, #a855f7);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        font-family: 'Orbitron', monospace; font-size: 16px !important;
    }
    .stSidebar .stMarkdown h3 {
        color: #00d4aa !important;
        border-left: 3px solid #00d4aa;
        padding-left: 8px;
        font-size: 13px !important;
    }

    /* ── Sidebar widgets ── */
    .stSidebar .stSlider > div > div > div {
        background: linear-gradient(90deg, #00bfff, #a855f7) !important;
    }
    .stSidebar .stSelectbox > div > div {
        background: #0d2137 !important;
        border: 1px solid #00bfff44 !important;
        border-radius: 8px !important;
        color: #e0e0e0 !important;
    }

    /* ── Metric Cards — floating glass ── */
    [data-testid="stMetric"] {
        background: rgba(255,255,255,0.04);
        border: 1px solid rgba(99,102,241,0.25);
        border-radius: 16px;
        padding: 12px 16px;
        box-shadow: 0 8px 32px rgba(0,0,0,0.4), inset 0 1px 0 rgba(255,255,255,0.08);
        -webkit-backdrop-filter: blur(12px);
        backdrop-filter: blur(12px);
        transition: all 0.3s cubic-bezier(0.4,0,0.2,1);
        animation: floatCard 4s ease-in-out infinite alternate;
    }
    [data-testid="stMetric"]:nth-child(odd)  { animation-delay: 0s; }
    [data-testid="stMetric"]:nth-child(even) { animation-delay: 0.5s; }
    @keyframes floatCard {
        0%   { transform: translateY(0px);   box-shadow: 0 8px 32px rgba(0,0,0,0.4); }
        100% { transform: translateY(-4px);  box-shadow: 0 16px 40px rgba(99,102,241,0.15); }
    }
    [data-testid="stMetric"]:hover {
        border-color: rgba(139,92,246,0.5);
        background: rgba(99,102,241,0.1);
        box-shadow: 0 16px 48px rgba(99,102,241,0.25), inset 0 1px 0 rgba(255,255,255,0.12);
        transform: translateY(-6px) !important;
    }
    [data-testid="stMetricValue"] {
        color: #ffffff !important;
        font-size: 1.3rem !important;
        font-family: 'Orbitron', monospace !important;
        text-shadow: 0 0 10px #00bfff88;
    }
    [data-testid="stMetricLabel"] {
        color: #00bfff !important;
        font-family: 'Rajdhani', sans-serif !important;
        font-weight: 600 !important;
        letter-spacing: 1px;
    }

    /* ── Headings ── */
    h1 {
        font-family: 'Orbitron', monospace !important;
        background: linear-gradient(90deg, #00bfff 0%, #a855f7 50%, #00d4aa 100%);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        font-size: 2rem !important;
        text-shadow: none;
        letter-spacing: 2px;
        animation: titleGlow 3s ease-in-out infinite alternate;
    }
    @keyframes titleGlow {
        from { filter: drop-shadow(0 0 8px #00bfff88); }
        to   { filter: drop-shadow(0 0 20px #a855f788); }
    }
    h2, h3 {
        font-family: 'Rajdhani', sans-serif !important;
        background: linear-gradient(90deg, #00d4aa, #00bfff);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        letter-spacing: 1px;
    }

    /* ── Feature Badges ── */
    .feature-badge {
        background: linear-gradient(135deg, #00bfff15, #a855f715);
        border: 1px solid;
        border-image: linear-gradient(90deg, #00bfff, #a855f7) 1;
        border-radius: 20px !important;
        padding: 5px 14px;
        font-size: 12px;
        font-family: 'Rajdhani', sans-serif;
        font-weight: 600;
        letter-spacing: 1px;
        display: inline-block; margin: 3px;
        background-clip: padding-box;
        color: #00d4aa;
        box-shadow: 0 0 10px #00bfff22;
        transition: all 0.3s;
    }
    .feature-badge:hover { box-shadow: 0 0 20px #00bfff55; }

    /* ── Safe / Unsafe boxes ── */
    .safe-box {
        background: linear-gradient(135deg, #00c85318, #00d4aa12);
        border: 2px solid #00c853;
        border-radius: 14px; padding: 14px 22px;
        font-size: 16px; font-weight: bold; color: #00c853;
        font-family: 'Rajdhani', sans-serif;
        box-shadow: 0 0 20px #00c85333, inset 0 1px 0 #00c85344;
        animation: safePulse 2s ease-in-out infinite;
    }
    @keyframes safePulse {
        0%,100% { box-shadow: 0 0 20px #00c85333; }
        50%      { box-shadow: 0 0 35px #00c85366; }
    }
    .unsafe-box {
        background: linear-gradient(135deg, #ff475718, #ff6b3512);
        border: 2px solid #ff4757;
        border-radius: 14px; padding: 14px 22px;
        font-size: 16px; font-weight: bold; color: #ff4757;
        font-family: 'Rajdhani', sans-serif;
        box-shadow: 0 0 20px #ff475733;
        animation: unsafePulse 1s ease-in-out infinite;
    }
    @keyframes unsafePulse {
        0%,100% { box-shadow: 0 0 20px #ff475744; }
        50%      { box-shadow: 0 0 40px #ff475799; }
    }

    /* ── AI Box ── */
    .ai-box {
        background: linear-gradient(135deg, #a855f715, #7c4dff12, #00bfff08);
        border: 1.5px solid;
        border-image: linear-gradient(135deg, #a855f7, #7c4dff, #00bfff) 1;
        border-radius: 14px; padding: 18px 22px; margin: 8px 0;
        box-shadow: 0 0 25px #7c4dff22;
        background-clip: padding-box;
    }
    .ai-title {
        font-family: 'Orbitron', monospace;
        background: linear-gradient(90deg, #a855f7, #7c4dff, #00bfff);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        font-size: 16px; font-weight: bold; margin-bottom: 12px;
        letter-spacing: 1px;
    }
    .ai-suggestion { color: #c8d8e8; font-size: 14px; padding: 4px 0; font-family: 'Rajdhani'; }
    .ai-warning {
        color: #ff6b6b; font-size: 14px; font-weight: bold; padding: 4px 0;
        text-shadow: 0 0 8px #ff4757aa;
    }
    .ai-good {
        color: #00e676; font-size: 14px; padding: 4px 0;
        text-shadow: 0 0 8px #00c85388;
    }

    /* ── Dynamic box ── */
    .dynamic-box {
        background: linear-gradient(135deg, #ff670015, #ffd70010);
        border: 1.5px solid #ff6700;
        border-radius: 12px; padding: 12px 18px; margin: 6px 0;
        box-shadow: 0 0 15px #ff670033;
    }

    /* ── Divider ── */
    hr {
        border: none;
        height: 1px;
        background: linear-gradient(90deg, transparent, #00bfff66, #a855f766, #00d4aa66, transparent);
        margin: 20px 0;
    }

    /* ── Buttons ── */
    .stButton > button, .stDownloadButton > button {
        background: linear-gradient(135deg, #0d2137, #1b3a5c) !important;
        border: 1.5px solid #00bfff66 !important;
        color: #00bfff !important;
        border-radius: 10px !important;
        font-family: 'Rajdhani', sans-serif !important;
        font-weight: 600 !important;
        letter-spacing: 1px !important;
        transition: all 0.3s !important;
        box-shadow: 0 0 10px #00bfff22 !important;
    }
    .stButton > button:hover, .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #00bfff22, #a855f722) !important;
        border-color: #00bfff !important;
        box-shadow: 0 0 20px #00bfff55 !important;
        transform: translateY(-1px) !important;
    }

    /* ── Expander ── */
    .streamlit-expanderHeader {
        background: linear-gradient(90deg, #0d2137, #1b3a5c) !important;
        border: 1px solid #00bfff33 !important;
        border-radius: 10px !important;
        color: #00bfff !important;
    }
    .streamlit-expanderContent {
        background: #080e1d !important;
        border: 1px solid #00bfff22 !important;
        border-top: none !important;
    }

    /* ── Dataframe ── */
    .stDataFrame { border: 1px solid #00bfff33 !important; border-radius: 10px; }

    /* ── Section divider glow line ── */
    .section-title {
        background: linear-gradient(90deg, #00bfff, #a855f7, #00d4aa);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        font-family: 'Orbitron', monospace;
        font-size: 18px; font-weight: 700;
        letter-spacing: 2px;
        display: inline-block;
    }

    /* ── Zoom Controls — cross-platform ── */
    .zoom-bar {
        position: fixed; bottom: 24px; right: 20px; z-index: 99999;
        display: flex; flex-direction: column; align-items: center; gap: 6px;
        background: linear-gradient(180deg, #050f20ee, #0a1628ee);
        border: 1.5px solid #00bfff44;
        border-radius: 16px; padding: 12px 10px;
        box-shadow: 0 0 30px #00bfff22, 0 8px 32px #00000099;
        -webkit-backdrop-filter: blur(10px);
        backdrop-filter: blur(10px);
        touch-action: none;
    }
    .zoom-btn {
        background: linear-gradient(135deg, #0d2137, #1b3a5c);
        border: 1.5px solid #00bfff66;
        color: #00bfff; font-size: 22px; font-weight: bold;
        width: 46px; height: 46px; border-radius: 12px;
        cursor: pointer;
        display: flex; align-items: center; justify-content: center;
        transition: all 0.15s;
        -webkit-user-select: none; user-select: none;
        -webkit-tap-highlight-color: transparent;
        box-shadow: 0 0 10px #00bfff22;
        /* Make tap targets big enough for mobile */
        min-width: 46px; min-height: 46px;
    }
    .zoom-btn:hover  { background: linear-gradient(135deg,#00bfff33,#a855f733); border-color:#00bfff; box-shadow:0 0 20px #00bfff55; }
    .zoom-btn:active { transform: scale(0.88); }
    .zoom-label {
        color: #00d4aa; font-size: 12px; font-weight: bold;
        font-family: 'Orbitron', monospace;
        text-align: center; letter-spacing: 1px;
        text-shadow: 0 0 8px #00d4aa88;
    }
    .zoom-reset {
        background: linear-gradient(135deg, #7c4dff22, #a855f722);
        border: 1.5px solid #a855f766;
        color: #a855f7; font-size: 10px; font-weight: bold;
        width: 46px; height: 30px; border-radius: 10px;
        cursor: pointer;
        display: flex; align-items: center; justify-content: center;
        transition: all 0.15s;
        -webkit-user-select: none; user-select: none;
        -webkit-tap-highlight-color: transparent;
        letter-spacing: 1px;
    }
    .zoom-reset:hover  { background: linear-gradient(135deg,#7c4dff44,#a855f744); border-color:#a855f7; }
    .zoom-reset:active { transform: scale(0.88); }
</style>

<div class="zoom-bar" id="zoomBar">
    <button class="zoom-btn" id="btnIn"  title="Zoom In">＋</button>
    <div class="zoom-label" id="zoom-pct">100%</div>
    <button class="zoom-btn" id="btnOut" title="Zoom Out">－</button>
    <button class="zoom-reset" id="btnReset">RESET</button>
</div>

<script>
(function(){
    var scale = 1.0;
    var MIN = 0.4, MAX = 2.8, STEP = 0.12;

    /* Find Streamlit main content — works on all platforms */
    function getTarget() {
        return document.querySelector('.main .block-container')
            || document.querySelector('[data-testid="block-container"]')
            || document.querySelector('.block-container')
            || document.querySelector('.main');
    }

    function applyZoom() {
        var el = getTarget();
        if (el) {
            el.style.transformOrigin = 'top center';
            el.style.transform = 'scale(' + scale.toFixed(2) + ')';
            /* Keep layout from collapsing on zoom out */
            if (scale < 1) {
                el.style.marginBottom = ((scale - 1) * el.scrollHeight) + 'px';
            } else {
                el.style.marginBottom = '';
            }
        }
        var lbl = document.getElementById('zoom-pct');
        if (lbl) lbl.textContent = Math.round(scale * 100) + '%';
        try { localStorage.setItem('beamZoomV2', scale.toFixed(2)); } catch(e){}
    }

    function zoomIn()    { scale = Math.min(MAX, parseFloat((scale + STEP).toFixed(2))); applyZoom(); }
    function zoomOut()   { scale = Math.max(MIN, parseFloat((scale - STEP).toFixed(2))); applyZoom(); }
    function zoomReset() { scale = 1.0; applyZoom(); }

    /* Restore saved zoom */
    try {
        var s = parseFloat(localStorage.getItem('beamZoomV2'));
        if (s && s >= MIN && s <= MAX) scale = s;
    } catch(e){}

    /* Button wiring — works on mouse + touch */
    function wire(id, fn) {
        var btn = document.getElementById(id);
        if (!btn) return;
        /* Touch (Android/iOS) */
        btn.addEventListener('touchstart', function(e){ e.preventDefault(); fn(); }, {passive:false});
        /* Mouse (Windows/Mac) */
        btn.addEventListener('mousedown', function(e){ e.preventDefault(); fn(); });
    }

    /* Keyboard shortcut: Ctrl/Cmd + / - */
    document.addEventListener('keydown', function(e){
        if (e.ctrlKey || e.metaKey) {
            if (e.key === '=' || e.key === '+') { e.preventDefault(); zoomIn(); }
            if (e.key === '-')                  { e.preventDefault(); zoomOut(); }
            if (e.key === '0')                  { e.preventDefault(); zoomReset(); }
        }
    });

    /* Init after Streamlit renders */
    function init() {
        wire('btnIn',    zoomIn);
        wire('btnOut',   zoomOut);
        wire('btnReset', zoomReset);
        applyZoom();
    }

    /* Multiple retries so Streamlit's dynamic DOM is ready */
    setTimeout(init, 600);
    setTimeout(applyZoom, 1200);
    setTimeout(applyZoom, 2500);

    /* Re-apply on any Streamlit re-render */
    var observer = new MutationObserver(function(){ applyZoom(); });
    setTimeout(function(){
        var root = document.querySelector('.main') || document.body;
        observer.observe(root, {childList:true, subtree:false});
    }, 1000);
})();
</script>
""", unsafe_allow_html=True)

st.markdown("# 🏗️ 2D Beam Stress & Deflection Simulator Pro")
st.markdown("""
<div style="margin: 8px 0 4px 0;">
    <span class="feature-badge">✨ 3D View</span>
    <span class="feature-badge">🏃 Dynamic Loads</span>
    <span class="feature-badge">📊 Excel Export</span>
    <span class="feature-badge">🧱 10 Materials</span>
    <span class="feature-badge">📐 Multi-Units</span>
    <span class="feature-badge">🔍 Zoom</span>
    <span class="feature-badge">🎨 Pro UI</span>
</div>
""", unsafe_allow_html=True)
st.markdown("---")

# ════════════════════════════════════════════
# SIDEBAR
# ════════════════════════════════════════════
st.sidebar.markdown("## ⚙️ Beam Parameters")
beam_type = st.sidebar.selectbox("🏛️ Beam Type", ["Simply Supported","Cantilever","Fixed-Fixed"])

# ── Unit System ──
st.sidebar.markdown("### 📐 Unit System")
length_unit = st.sidebar.selectbox("Length Unit", ["m", "cm", "mm"], index=0)
force_unit  = st.sidebar.selectbox("Force Unit",  ["N", "kN"], index=0)

# Conversion factors → always compute in SI (m, N)
L_CONV = {"m": 1.0, "cm": 0.01, "mm": 0.001}[length_unit]
F_CONV = {"N": 1.0, "kN": 1000.0}[force_unit]

# Max length in display unit
L_max_display = {"m": 100.0, "cm": 10000.0, "mm": 100000.0}[length_unit]
L_default     = {"m": 4.0,   "cm": 400.0,   "mm": 4000.0}[length_unit]
L_min         = {"m": 0.1,   "cm": 10.0,    "mm": 100.0}[length_unit]
L_step        = {"m": 0.1,   "cm": 10.0,    "mm": 100.0}[length_unit]

L_display = st.sidebar.slider(f"📏 Beam Length ({length_unit})", L_min, L_max_display, L_default, L_step)
L = L_display * L_CONV  # always in metres internally

st.sidebar.markdown("### ⬇️ Point Loads")
n_loads = st.sidebar.number_input("Number of Point Loads", 1, 8, 1)

# Force slider max in display unit
F_max = {"N": 10000, "kN": 100}[force_unit]
F_def = {"N": 500,   "kN": 5}[force_unit]
F_step= {"N": 100,   "kN": 1}[force_unit]

# Position slider max in display unit
pos_max = L_display

loads = []
for i in range(int(n_loads)):
    st.sidebar.markdown(f"**Load {i+1}**")
    p_disp = st.sidebar.slider(f"P{i+1} ({force_unit})", 0, F_max, F_def, F_step, key=f"p{i}")
    default_pos = round(L_display/(n_loads+1)*(i+1), 1)
    default_pos = min(default_pos, L_display)
    a_disp = st.sidebar.slider(f"Position {i+1} ({length_unit})", 0.0, L_display, default_pos, L_step, key=f"a{i}")
    loads.append((p_disp * F_CONV, a_disp * L_CONV))  # store in SI

st.sidebar.markdown("### 〰️ UDL")
udl_unit_label = f"{force_unit}/{length_unit}"
w_max  = {"N": 2000, "kN": 20}[force_unit]
w_step = {"N": 50,   "kN": 1}[force_unit]
w_disp = st.sidebar.slider(f"UDL Intensity ({udl_unit_label})", 0, w_max, 0, w_step)
w = w_disp * F_CONV / L_CONV  # convert to N/m internally

# ── NEW: Dynamic/Moving Load ──
st.sidebar.markdown("### 🚗 Dynamic / Moving Load")
enable_dynamic = st.sidebar.checkbox("Enable Moving Load Analysis", False)
if enable_dynamic:
    P_moving = st.sidebar.slider("Moving Load (N)", 100, 5000, 1000, 100)
    dynamic_positions = 20

# ── NEW: Extended Material Library ──
material_data = {
    "🔩 Structural Steel (200 GPa)":    {"E": 200e9, "yield": 250e6,  "density": 7850, "color": "#4fc3f7"},
    "🔩 High-Strength Steel (200 GPa)": {"E": 200e9, "yield": 690e6,  "density": 7850, "color": "#0288d1"},
    "✈️ Aluminum 6061 (69 GPa)":        {"E": 69e9,  "yield": 276e6,  "density": 2700, "color": "#b0bec5"},
    "✈️ Aluminum 7075 (72 GPa)":        {"E": 72e9,  "yield": 503e6,  "density": 2810, "color": "#90a4ae"},
    "🌳 Oak Wood (12 GPa)":             {"E": 12e9,  "yield": 40e6,   "density": 700,  "color": "#8d6e63"},
    "🌲 Pine Wood (9 GPa)":             {"E": 9e9,   "yield": 33e6,   "density": 550,  "color": "#a1887f"},
    "🏗️ Concrete (30 GPa)":            {"E": 30e9,  "yield": 30e6,   "density": 2400, "color": "#9e9e9e"},
    "⚡ Copper (110 GPa)":             {"E": 110e9, "yield": 210e6,  "density": 8900, "color": "#ff8f00"},
    "🔮 Carbon Fiber (150 GPa)":        {"E": 150e9, "yield": 600e6,  "density": 1600, "color": "#7c4dff"},
    "🪨 Titanium (116 GPa)":           {"E": 116e9, "yield": 880e6,  "density": 4500, "color": "#80cbc4"},
}
E_sel = st.sidebar.selectbox("🧱 Material", list(material_data.keys()))
E_val    = material_data[E_sel]["E"]
yield_s  = material_data[E_sel]["yield"]
density  = material_data[E_sel]["density"]
mat_color = material_data[E_sel]["color"]

st.sidebar.markdown("### 📐 Cross Section Type")
section_type = st.sidebar.selectbox("Section Shape", ["Rectangular","Circular","I-Beam","T-Beam"])

def get_section_properties(section_type, sidebar):
    if section_type == "Rectangular":
        b = sidebar.number_input("Width b (m)",  0.01, 0.5, 0.05)
        h = sidebar.number_input("Height h (m)", 0.01, 0.5, 0.10)
        return (b*h**3)/12, h/2, h, b, {"b":b,"h":h}
    elif section_type == "Circular":
        d = sidebar.number_input("Diameter d (m)", 0.01, 0.5, 0.08)
        return (np.pi*d**4)/64, d/2, d, d, {"d":d}
    elif section_type == "I-Beam":
        sidebar.markdown("*Flange & Web*")
        bf = sidebar.number_input("Flange Width bf (m)", 0.02, 0.4, 0.10)
        tf = sidebar.number_input("Flange Thick tf (m)", 0.005, 0.05, 0.01)
        hw = sidebar.number_input("Web Height hw (m)",   0.05, 0.4,  0.10)
        tw = sidebar.number_input("Web Thick tw (m)",    0.005, 0.05, 0.006)
        ht = hw+2*tf
        I  = (bf*ht**3)/12 - ((bf-tw)*hw**3)/12
        return I, ht/2, ht, bf, {"bf":bf,"tf":tf,"hw":hw,"tw":tw}
    elif section_type == "T-Beam":
        sidebar.markdown("*Flange & Web*")
        bf = sidebar.number_input("Flange Width bf (m)", 0.02, 0.4, 0.10)
        tf = sidebar.number_input("Flange Thick tf (m)", 0.005, 0.05, 0.01)
        hw = sidebar.number_input("Web Height hw (m)",   0.05, 0.4,  0.10)
        tw = sidebar.number_input("Web Thick tw (m)",    0.005, 0.05, 0.006)
        ht = hw+tf
        Af = bf*tf; Aw = tw*hw; At = Af+Aw
        yb = (Af*(ht-tf/2)+Aw*(hw/2))/At
        I  = (bf*tf**3)/12+Af*(ht-tf/2-yb)**2+(tw*hw**3)/12+Aw*(hw/2-yb)**2
        return I, max(yb,ht-yb), ht, bf, {"bf":bf,"tf":tf,"hw":hw,"tw":tw,"y_bar":yb}

I, c_dist, h_total, b_total, dims = get_section_properties(section_type, st.sidebar)

# ════════════════════════════════════════════
# CALCULATIONS
# ════════════════════════════════════════════
x = np.linspace(0, L, 2000)
colors_load = ['#ff4757','#ffd700','#ff6b35','#7c4dff','#00d4aa','#00bfff','#ff69b4','#adff2f']

if beam_type == "Simply Supported":
    RA = sum((P*(L-a))/L for P,a in loads) + w*L/2
    RB = sum((P*a)/L     for P,a in loads) + w*L/2
    V  = np.full_like(x, RA)
    for P,a in loads: V -= P*(x >= a).astype(float)
    V -= w*x
    M  = RA*x - 0.5*w*x**2
    for P,a in loads: M -= P*np.maximum(x-a, 0)
    y  = (w*x)/(24*E_val*I)*(L**3-2*L*x**2+x**3)
    for P,a in loads:
        bv = L-a
        y += np.where(x<a,
            (P*bv*x)/(6*E_val*I*L)*(L**2-bv**2-x**2),
            (P*a*(L-x))/(6*E_val*I*L)*(2*L*(L-x)-a**2-(L-x)**2))

elif beam_type == "Cantilever":
    RA = sum(P for P,a in loads) + w*L
    MA_fix = sum(P*(L-a) for P,a in loads) + w*L**2/2
    RB = 0
    V  = np.full_like(x, RA)
    for P,a in loads: V -= P*(x >= a).astype(float)
    V -= w*x
    M  = -MA_fix + RA*x - 0.5*w*x**2
    for P,a in loads: M -= P*np.maximum(x-a, 0)
    y  = np.zeros_like(x)
    for P,a in loads:
        y += np.where(x<=a,
            (P*x**2)/(6*E_val*I)*(3*a-x),
            (P*a**2)/(6*E_val*I)*(3*x-a))
    y += (w*x**2)/(24*E_val*I)*(x**2-4*L*x+6*L**2)
    y  = y - y[0]

elif beam_type == "Fixed-Fixed":
    RA = sum((P*(L-a)**2*(2*a+L))/L**3 for P,a in loads) + w*L/2
    RB = sum((P*a**2*(3*L-2*a))/L**3   for P,a in loads) + w*L/2
    MA_fix = sum((P*a*(L-a)**2)/L**2   for P,a in loads) + w*L**2/12
    V  = np.full_like(x, RA)
    for P,a in loads: V -= P*(x >= a).astype(float)
    V -= w*x
    M  = -MA_fix + RA*x - 0.5*w*x**2
    for P,a in loads: M -= P*np.maximum(x-a, 0)
    y  = np.zeros_like(x)
    for P,a in loads:
        bv = L-a
        y += np.where(x<a,
            (P*bv**2*x**2)/(6*E_val*I*L**3)*(3*a*L-x*(3*a+bv)),
            (P*a**2*(L-x)**2)/(6*E_val*I*L**3)*(3*bv*L-(L-x)*(3*bv+a)))
    y += (w*x**2*(L-x)**2)/(24*E_val*I)

M_max     = max(abs(M))
y_max     = max(abs(y))*1000
sigma_max = (M_max*c_dist)/I/1e6
yield_MPa = yield_s/1e6
FOS       = yield_MPa/sigma_max if sigma_max > 0 else 999

# ── Weight Calculation ──
if section_type == "Rectangular":
    area = dims['b']*dims['h']
elif section_type == "Circular":
    area = np.pi*(dims['d']/2)**2
elif section_type in ["I-Beam","T-Beam"]:
    area = I / (c_dist**2) * 1.2  # approximate
else:
    area = 0.01
beam_weight = density * area * L * 9.81  # N

# ════════════════════════════════════════════
# AI SUGGESTIONS
# ════════════════════════════════════════════
def ai_suggestions(beam_type,section_type,L,loads,w,E,dims,
                   FOS,sigma_max,yield_MPa,y_max,M_max,I,h_total):
    suggestions=[]; warnings=[]; good=[]
    if FOS < 1.0:
        warnings.append("🚨 CRITICAL: Beam will fail immediately! Redesign required.")
    elif FOS < 1.5:
        warnings.append(f"⚠️ FOS dangerously low ({FOS:.2f}). Minimum recommended: 2.0")
    elif FOS < 2.0:
        warnings.append(f"⚠️ FOS ({FOS:.2f}) below recommended minimum of 2.0")
    else:
        good.append(f"✅ Factor of Safety is adequate ({FOS:.2f} ≥ 2.0)")
    stress_ratio = sigma_max/yield_MPa if yield_MPa > 0 else 0
    if stress_ratio > 0.9:
        warnings.append(f"🔴 Max stress ({sigma_max:.1f} MPa) is {stress_ratio*100:.0f}% of yield — critical!")
    elif stress_ratio > 0.6:
        warnings.append(f"🟡 Max stress is {stress_ratio*100:.0f}% of yield — moderate risk")
    else:
        good.append(f"✅ Stress level is safe ({stress_ratio*100:.0f}% of yield strength)")
    if y_max > 0:
        defl_ratio = (y_max/1000)/L
        if defl_ratio > 1/200:
            warnings.append(f"⚠️ Excessive deflection: L/{int(L*1000/y_max)} (recommended ≤ L/200)")
            suggestions.append("💡 Increase section height to reduce deflection")
        else:
            good.append(f"✅ Deflection within limit: L/{int(L*1000/y_max)}")
    if section_type == "Rectangular":
        h = dims.get('h',0.1)
        if FOS < 2.0:
            new_h = h*(2.0/FOS)**0.5
            suggestions.append(f"💡 Increase height from {h*1000:.0f}mm to {new_h*1000:.0f}mm for FOS ≥ 2.0")
    elif section_type == "Circular":
        d = dims.get('d',0.08)
        if FOS < 2.0:
            new_d = d*(2.0/FOS)**0.333
            suggestions.append(f"💡 Increase diameter from {d*1000:.0f}mm to {new_d*1000:.0f}mm for FOS ≥ 2.0")
    elif section_type == "I-Beam":
        good.append("✅ I-Beam is efficient for bending — good section choice!")
    elif section_type == "T-Beam":
        good.append("✅ T-Beam is good for one-directional bending")
    if "Steel" in E and FOS > 10:
        suggestions.append("💡 FOS very high — consider Aluminum to reduce weight by ~65%")
    if "Aluminum" in E and FOS < 2.0:
        suggestions.append("💡 Switch to Steel for 3x higher yield strength")
    if "Wood" in E or "Pine" in E or "Oak" in E:
        suggestions.append("💡 Wood is orthotropic — verify grain direction matches load direction")
    if "Concrete" in E:
        warnings.append("⚠️ Concrete is weak in tension — add steel reinforcement (rebar)!")
    if "Carbon Fiber" in E:
        good.append("✅ Carbon fiber: excellent strength-to-weight ratio for aerospace/high-perf apps")
    if beam_type == "Simply Supported" and FOS < 2.0:
        suggestions.append("💡 Consider Fixed-Fixed beam — reduces max moment by ~50%")
    if beam_type == "Cantilever" and FOS < 2.0:
        suggestions.append("💡 Cantilever has high moment at fixed end — add intermediate support")
    if len(loads)==1 and loads[0][0]>0:
        if abs(loads[0][1]-L/2) < 0.1:
            good.append("✅ Load at center — symmetric loading condition")
    if FOS >= 3.0 and stress_ratio < 0.4:
        good.append("🏆 Excellent design! Well within all safety limits.")
    elif FOS >= 2.0:
        good.append("👍 Good design — meets standard engineering requirements.")
    return warnings, suggestions, good

# ════════════════════════════════════════════
# BEAM VISUAL
# ════════════════════════════════════════════
st.markdown("### 🔎 Beam Diagram")
fig_beam, ax_b = plt.subplots(figsize=(10, 2.8))
fig_beam.patch.set_facecolor('#0a1628')
ax_b.set_facecolor('#0d2137')
ax_b.set_xlim(-0.8, L+0.8); ax_b.set_ylim(-1.2, 2.8); ax_b.axis('off')
ax_b.add_patch(plt.Rectangle((0,0.2),L,0.3,color='#1b3a5c',zorder=2))
ax_b.add_patch(plt.Rectangle((0,0.2),L,0.3,fill=False,edgecolor='#00bfff',lw=2,zorder=3))
if beam_type == "Simply Supported":
    ax_b.plot([0],[0.2],'v',color='#00d4aa',markersize=14,zorder=4)
    ax_b.plot([L],[0.2],'o',color='#00d4aa',markersize=10,zorder=4)
    ax_b.text(0,-0.3,'A (Pin)',color='#00d4aa',ha='center',fontsize=8)
    ax_b.text(L,-0.3,'B (Roller)',color='#00d4aa',ha='center',fontsize=8)
elif beam_type == "Cantilever":
    ax_b.add_patch(plt.Rectangle((-0.3,-0.1),0.3,0.9,color='#00d4aa',zorder=4))
    ax_b.text(0,-0.3,'Fixed',color='#00d4aa',ha='center',fontsize=8)
    ax_b.text(L,-0.3,'Free End',color='#ff4757',ha='center',fontsize=8)
elif beam_type == "Fixed-Fixed":
    ax_b.add_patch(plt.Rectangle((-0.3,-0.1),0.3,0.9,color='#00d4aa',zorder=4))
    ax_b.add_patch(plt.Rectangle((L,-0.1),0.3,0.9,color='#00d4aa',zorder=4))
    ax_b.text(0,-0.3,'Fixed A',color='#00d4aa',ha='center',fontsize=8)
    ax_b.text(L,-0.3,'Fixed B',color='#00d4aa',ha='center',fontsize=8)
for i,(P,a) in enumerate(loads):
    if P > 0:
        p_disp_label = P / F_CONV
        ax_b.annotate('',xy=(a,0.5),xytext=(a,1.6),
            arrowprops=dict(arrowstyle='->',color=colors_load[i],lw=2.5))
        ax_b.text(a,1.75,f'P{i+1}={p_disp_label:.0f}{force_unit}',color=colors_load[i],ha='center',fontsize=8,fontweight='bold')
if w > 0:
    w_disp_label = w * L_CONV / F_CONV
    for xi in np.linspace(0.1,L-0.1,14):
        ax_b.annotate('',xy=(xi,0.5),xytext=(xi,1.1),
            arrowprops=dict(arrowstyle='->',color='#00bfff',lw=1))
    ax_b.plot([0,L],[1.15,1.15],color='#00bfff',lw=2)
    ax_b.text(L/2,1.28,f'w={w_disp_label:.1f} {force_unit}/{length_unit}',color='#00bfff',ha='center',fontsize=8)
ax_b.annotate('',xy=(L,-0.7),xytext=(0,-0.7),
    arrowprops=dict(arrowstyle='<->',color='gray',lw=1))
ax_b.text(L/2,-1.0,f'L = {L_display}{length_unit}',color='gray',ha='center',fontsize=8)
st.pyplot(fig_beam)
st.markdown("---")

# ════════════════════════════════════════════
# SAFETY CHECK
# ════════════════════════════════════════════
st.markdown("### 🛡️ Safety Check")
if sigma_max < yield_MPa:
    st.markdown(f"""<div class='safe-box'>✅ SAFE &nbsp;|&nbsp; σ_max: {sigma_max:.2f} MPa &nbsp;<&nbsp; Yield: {yield_MPa:.0f} MPa &nbsp;|&nbsp; FOS: {FOS:.2f} &nbsp;|&nbsp; Weight: {beam_weight:.1f} N</div>""", unsafe_allow_html=True)
else:
    st.markdown(f"""<div class='unsafe-box'>❌ UNSAFE &nbsp;|&nbsp; σ_max: {sigma_max:.2f} MPa &nbsp;>&nbsp; Yield: {yield_MPa:.0f} MPa &nbsp;|&nbsp; BEAM WILL FAIL!</div>""", unsafe_allow_html=True)
st.markdown("---")

# ════════════════════════════════════════════
# RESULTS
# ════════════════════════════════════════════
st.markdown("### 📊 Results")
c1,c2,c3,c4,c5,c6,c7,c8 = st.columns(8)
c1.metric("RA", f"{RA:.1f} N")
c2.metric("RB", f"{RB:.1f} N")
c3.metric("M_max", f"{M_max:.1f} N·m")
c4.metric("y_max", f"{y_max:.3f} mm")
c5.metric("σ_max", f"{sigma_max:.2f} MPa")
c6.metric("FOS",   f"{FOS:.2f}")
c7.metric("I",     f"{I:.2e} m⁴")
c8.metric("Weight",f"{beam_weight:.0f} N")
st.markdown("---")

# ════════════════════════════════════════════
# AI SUGGESTIONS
# ════════════════════════════════════════════
st.markdown("### 🤖 AI Engineering Suggestions")
warnings, suggestions, good = ai_suggestions(
    beam_type,section_type,L,loads,w,E_sel,dims,
    FOS,sigma_max,yield_MPa,y_max,M_max,I,h_total)
ai_html = "<div class='ai-box'><div class='ai-title'>🤖 Smart Engineering Analysis</div>"
for warn in warnings: ai_html += f"<div class='ai-warning'>{warn}</div>"
for sugg in suggestions: ai_html += f"<div class='ai-suggestion'>{sugg}</div>"
for g in good: ai_html += f"<div class='ai-good'>{g}</div>"
ai_html += "</div>"
st.markdown(ai_html, unsafe_allow_html=True)
st.markdown("---")

# ════════════════════════════════════════════
# PLOTS (SFD, BMD, Deflection)
# ════════════════════════════════════════════
st.markdown("### 📈 Engineering Diagrams")
fig,(ax1,ax2,ax3) = plt.subplots(3,1,figsize=(10,9))
fig.patch.set_facecolor('#0a1628')
for ax in [ax1,ax2,ax3]:
    ax.set_facecolor('#0d2137'); ax.tick_params(colors='#aaaaaa')
    ax.spines[:].set_color('#1b3a5c')
    ax.yaxis.label.set_color('#aaaaaa'); ax.xaxis.label.set_color('#aaaaaa')
    ax.title.set_color('#00bfff'); ax.grid(True,color='#1b3a5c',linewidth=0.7)
w_disp_label = w * L_CONV / F_CONV if w > 0 else 0
fig.suptitle(f"{beam_type} | {section_type} | L={L_display}{length_unit} | UDL={w_disp_label:.1f}{force_unit}/{length_unit}",
             fontsize=11,fontweight='bold',color='white')
ax1.plot(x,V,color='#00bfff',lw=2); ax1.fill_between(x,V,alpha=0.25,color='#00bfff')
ax1.axhline(0,color='white',lw=0.8)
for i,(P,a) in enumerate(loads):
    ax1.axvline(a,color=colors_load[i],lw=1,linestyle='--',alpha=0.7,label=f'P{i+1}@{a/L_CONV:.1f}{length_unit}')
ax1.set_ylabel(f"Shear Force ({force_unit})"); ax1.set_title("SFD")
ax1.legend(facecolor='#0d2137',labelcolor='white',fontsize=7)
ax2.plot(x,M,color='#ff6b35',lw=2); ax2.fill_between(x,M,alpha=0.25,color='#ff6b35')
ax2.axhline(0,color='white',lw=0.8)
ax2.set_ylabel(f"Bending Moment ({force_unit}·{length_unit})"); ax2.set_title("BMD")
ax3.plot(x,y*1000,color='#00d4aa',lw=2); ax3.fill_between(x,y*1000,alpha=0.25,color='#00d4aa')
ax3.axhline(0,color='white',lw=0.8)
ax3.set_xlabel(f"Position ({length_unit})"); ax3.set_ylabel("Deflection (mm)"); ax3.set_title("Deflection Curve")
plt.tight_layout()
st.pyplot(fig)

# ════════════════════════════════════════════
# NEW FEATURE 1: 3D BEAM VISUALIZATION
# ════════════════════════════════════════════
st.markdown("---")
st.markdown("### 🎯 3D Beam Visualization")

def plot_3d_beam():
    fig3d = plt.figure(figsize=(12, 5))
    fig3d.patch.set_facecolor('#0a1628')
    
    ax3d = fig3d.add_subplot(121, projection='3d')
    ax3d.set_facecolor('#0d2137')
    
    # Beam geometry
    n_x = 50
    x3 = np.linspace(0, L, n_x)
    
    # Cross-section width and height
    if section_type == "Rectangular":
        bw, bh = dims['b'], dims['h']
    elif section_type == "Circular":
        bw, bh = dims['d'], dims['d']
    else:
        bw = b_total
        bh = h_total
    
    # Deflection along beam
    y_defl = np.interp(x3, x, y)
    
    # Draw beam as 3D box with deflection
    X = np.array([x3, x3])
    Z = np.array([np.zeros(n_x)-bh/2, np.zeros(n_x)+bh/2])
    Y_top = np.array([y_defl, y_defl]) - bh/2
    Y_bot = np.array([y_defl, y_defl]) + bh/2
    
    # Top face
    ax3d.plot_surface(X, np.array([[-bw/2]*n_x, [bw/2]*n_x]),
                      np.array([y_defl, y_defl])*1000 + bh/2*500,
                      alpha=0.6, color=mat_color)
    # Bottom face
    ax3d.plot_surface(X, np.array([[-bw/2]*n_x, [bw/2]*n_x]),
                      np.array([y_defl, y_defl])*1000 - bh/2*500,
                      alpha=0.6, color='#1b3a5c')
    
    # Deflection line
    ax3d.plot(x3, [0]*n_x, y_defl*1000, color='#ff4757', lw=2, label='Neutral Axis')
    
    # Load arrows
    for i,(P,a) in enumerate(loads):
        if P > 0:
            y_at_a = np.interp(a, x, y)*1000
            ax3d.quiver(a, 0, y_at_a+200, 0, 0, -150,
                       color=colors_load[i], arrow_length_ratio=0.3, lw=2)
    
    ax3d.set_xlabel('Length (m)', color='#aaaaaa', fontsize=7)
    ax3d.set_ylabel('Width', color='#aaaaaa', fontsize=7)
    ax3d.set_zlabel('Deflection (mm)', color='#aaaaaa', fontsize=7)
    ax3d.set_title(f'3D Beam — {section_type}', color='#00bfff', fontsize=9)
    ax3d.tick_params(colors='#aaaaaa', labelsize=6)
    ax3d.xaxis.pane.fill = False
    ax3d.yaxis.pane.fill = False
    ax3d.zaxis.pane.fill = False
    ax3d.grid(True, color='#1b3a5c', alpha=0.3)
    
    # Stress color map on beam side view
    ax_stress = fig3d.add_subplot(122)
    ax_stress.set_facecolor('#0d2137')
    
    x_cs = np.linspace(0, L, 200)
    y_cs = np.linspace(-h_total/2, h_total/2, 60)
    M_cs = np.interp(x_cs, x, M)
    SIGMA_cs = np.outer(y_cs, M_cs)/I/1e6
    
    cmap3d = mcolors.LinearSegmentedColormap.from_list('stress3d',
        ['#0000ff','#00bfff','#00ff00','#ffff00','#ff0000'])
    vmax3d = sigma_max if sigma_max > 0 else 1
    im3d = ax_stress.imshow(SIGMA_cs, aspect='auto', origin='lower',
        extent=[0, L, -h_total/2*1000, h_total/2*1000],
        cmap=cmap3d, vmin=-vmax3d, vmax=vmax3d)
    
    # Plot deflected shape overlay
    ax_stress.plot(x, y*1000*5, color='white', lw=2, linestyle='--', label=f'Deflection ×5')
    ax_stress.legend(facecolor='#0d2137', labelcolor='white', fontsize=7)
    
    cbar3d = fig3d.colorbar(im3d, ax=ax_stress, pad=0.02)
    cbar3d.set_label('Stress (MPa)', color='white', fontsize=8)
    cbar3d.ax.yaxis.set_tick_params(color='white')
    plt.setp(cbar3d.ax.yaxis.get_ticklabels(), color='white')
    ax_stress.set_xlabel("Position (m)", color='#aaaaaa')
    ax_stress.set_ylabel("Height (mm)", color='#aaaaaa')
    ax_stress.set_title("Stress + Deflection Side View", color='#00bfff', fontsize=9)
    ax_stress.tick_params(colors='#aaaaaa')
    ax_stress.spines[:].set_color('#1b3a5c')
    
    plt.tight_layout()
    return fig3d

fig3d = plot_3d_beam()
st.pyplot(fig3d)

# ════════════════════════════════════════════
# NEW FEATURE 2: DYNAMIC / MOVING LOAD
# ════════════════════════════════════════════
st.markdown("---")
st.markdown("### 🚗 Dynamic / Moving Load Analysis")

if enable_dynamic:
    st.markdown(f"""<div class='dynamic-box'>
    🚗 <b>Moving Load: {P_moving} N</b> — Analyzing influence lines across beam length
    </div>""", unsafe_allow_html=True)
    
    pos_arr = np.linspace(0.01, L-0.01, dynamic_positions)
    M_env_max = np.zeros_like(x)
    M_env_min = np.zeros_like(x)
    y_env_max = np.zeros_like(x)
    
    all_M = []
    all_y_defl = []
    
    for pos in pos_arr:
        if beam_type == "Simply Supported":
            RA_d = P_moving*(L-pos)/L
            M_d  = RA_d*x - P_moving*np.maximum(x-pos, 0)
            y_d  = np.where(x < pos,
                (P_moving*(L-pos)*x)/(6*E_val*I*L)*(L**2-(L-pos)**2-x**2),
                (P_moving*pos*(L-x))/(6*E_val*I*L)*(2*L*(L-x)-pos**2-(L-x)**2))
        elif beam_type == "Cantilever":
            M_d = np.where(x <= pos, -P_moving*(pos-x), 0)
            y_d = np.where(x <= pos,
                (P_moving*x**2)/(6*E_val*I)*(3*pos-x),
                (P_moving*pos**2)/(6*E_val*I)*(3*x-pos))
        else:  # Fixed-Fixed
            bv = L - pos
            RA_d = P_moving*bv**2*(2*pos+L)/L**3
            MA_d = P_moving*pos*bv**2/L**2
            M_d  = -MA_d + RA_d*x - P_moving*np.maximum(x-pos, 0)
            y_d  = np.where(x < pos,
                (P_moving*bv**2*x**2)/(6*E_val*I*L**3)*(3*pos*L-x*(3*pos+bv)),
                (P_moving*pos**2*(L-x)**2)/(6*E_val*I*L**3)*(3*bv*L-(L-x)*(3*bv+pos)))
        
        all_M.append(M_d)
        all_y_defl.append(y_d*1000)
        M_env_max = np.maximum(M_env_max, M_d)
        M_env_min = np.minimum(M_env_min, M_d)
        y_env_max = np.maximum(y_env_max, np.abs(y_d*1000))
    
    fig_dyn, (ax_d1, ax_d2, ax_d3) = plt.subplots(3, 1, figsize=(10, 9))
    fig_dyn.patch.set_facecolor('#0a1628')
    for ax in [ax_d1, ax_d2, ax_d3]:
        ax.set_facecolor('#0d2137')
        ax.tick_params(colors='#aaaaaa')
        ax.spines[:].set_color('#1b3a5c')
        ax.grid(True, color='#1b3a5c', linewidth=0.5)
        ax.yaxis.label.set_color('#aaaaaa')
        ax.xaxis.label.set_color('#aaaaaa')
        ax.title.set_color('#00bfff')
    
    fig_dyn.suptitle(f"Moving Load Analysis — P={P_moving}N on {beam_type}", 
                     color='white', fontsize=11, fontweight='bold')
    
    # All BMD positions (influence lines)
    for i, M_d in enumerate(all_M):
        alpha = 0.15 + 0.5*(i/len(all_M))
        ax_d1.plot(x, M_d, color='#ff6b35', alpha=0.3, lw=0.8)
    ax_d1.fill_between(x, M_env_max, M_env_min, alpha=0.3, color='#ff6b35', label='Moment Envelope')
    ax_d1.plot(x, M_env_max, color='#ff4757', lw=2, label='Max BMD')
    ax_d1.plot(x, M_env_min, color='#ffd700', lw=2, label='Min BMD')
    ax_d1.axhline(0, color='white', lw=0.8)
    ax_d1.set_ylabel("Bending Moment (N·m)")
    ax_d1.set_title("BMD Envelope — All Load Positions")
    ax_d1.legend(facecolor='#0d2137', labelcolor='white', fontsize=7)
    
    # Deflection envelope
    for y_d in all_y_defl:
        ax_d2.plot(x, y_d, color='#00d4aa', alpha=0.2, lw=0.8)
    ax_d2.plot(x, y_env_max, color='#00d4aa', lw=2.5, label='Max Deflection Envelope')
    ax_d2.axhline(0, color='white', lw=0.8)
    ax_d2.set_ylabel("Deflection (mm)")
    ax_d2.set_title("Deflection Envelope")
    ax_d2.legend(facecolor='#0d2137', labelcolor='white', fontsize=7)
    
    # Max moment vs load position (influence line for center)
    mid_idx = len(x)//2
    M_at_mid = [M_d[mid_idx] for M_d in all_M]
    ax_d3.plot(pos_arr, M_at_mid, color='#7c4dff', lw=2.5, marker='o', markersize=4)
    ax_d3.fill_between(pos_arr, M_at_mid, alpha=0.25, color='#7c4dff')
    ax_d3.axhline(0, color='white', lw=0.8)
    ax_d3.set_xlabel("Load Position (m)")
    ax_d3.set_ylabel("Moment at Midspan (N·m)")
    ax_d3.set_title("Influence Line — Midspan Moment")
    
    plt.tight_layout()
    st.pyplot(fig_dyn)
    
    max_dyn_moment = np.max(M_env_max)
    max_dyn_defl   = np.max(y_env_max)
    st.info(f"🚗 Dynamic Analysis: Peak Moment = **{max_dyn_moment:.1f} N·m** | Peak Deflection = **{max_dyn_defl:.3f} mm** (worst case load position)")
else:
    st.info("☝️ Enable **Moving Load Analysis** in the sidebar to see influence lines & envelopes")

# ════════════════════════════════════════════
# STRESS HEATMAP
# ════════════════════════════════════════════
st.markdown("---")
st.markdown("### 🌡️ Stress Heatmap (ANSYS Style)")
x_h = np.linspace(0,L,300)
y_h = np.linspace(-h_total/2,h_total/2,80)
M_h = np.interp(x_h,x,M)
SIGMA = np.outer(y_h,M_h)/I/1e6
fig_hm,ax_hm = plt.subplots(figsize=(12,3))
fig_hm.patch.set_facecolor('#0a1628'); ax_hm.set_facecolor('#0a1628')
cmap = mcolors.LinearSegmentedColormap.from_list('ansys',
    ['#0000ff','#00bfff','#00ff00','#ffff00','#ff6600','#ff0000'])
vmax = sigma_max if sigma_max > 0 else 1
im = ax_hm.imshow(SIGMA,aspect='auto',origin='lower',
    extent=[0,L,-h_total/2*1000,h_total/2*1000],cmap=cmap,vmin=-vmax,vmax=vmax)
cbar = fig_hm.colorbar(im,ax=ax_hm,pad=0.02)
cbar.set_label('Stress (MPa)',color='white',fontsize=9)
cbar.ax.yaxis.set_tick_params(color='white')
plt.setp(cbar.ax.yaxis.get_ticklabels(),color='white')
ax_hm.set_xlabel("Position along beam (m)",color='#aaaaaa')
ax_hm.set_ylabel("Height (mm)",color='#aaaaaa')
ax_hm.set_title(f"Bending Stress Heatmap — {section_type} | {E_sel}",color='#00bfff',fontsize=10)
ax_hm.tick_params(colors='#aaaaaa'); ax_hm.spines[:].set_color('#1b3a5c')
st.pyplot(fig_hm)

# ════════════════════════════════════════════
# CROSS SECTION VISUALIZER
# ════════════════════════════════════════════
st.markdown("---")
st.markdown("### 📐 Cross Section Visualizer")
fig_cs,(ax_cs,ax_sd) = plt.subplots(1,2,figsize=(10,4))
fig_cs.patch.set_facecolor('#0a1628')
for ax in [ax_cs,ax_sd]:
    ax.set_facecolor('#0d2137'); ax.tick_params(colors='#aaaaaa')
    ax.spines[:].set_color('#1b3a5c'); ax.grid(True,color='#1b3a5c')
ax_cs.set_title(f"{section_type} Cross Section",color='#00bfff')
if section_type == "Rectangular":
    b,h = dims['b'],dims['h']
    ax_cs.add_patch(mpatches.Rectangle((-b/2,-h/2),b,h,color=mat_color+'44',ec='#00bfff',lw=2))
    ax_cs.set_xlim(-b,b); ax_cs.set_ylim(-h,h)
    ax_cs.axhline(0,color='#ffd700',lw=1.5,linestyle='--',label='Neutral Axis')
elif section_type == "Circular":
    d = dims['d']
    ax_cs.add_patch(mpatches.Circle((0,0),d/2,color=mat_color+'44',ec='#00bfff',lw=2))
    ax_cs.set_xlim(-d,d); ax_cs.set_ylim(-d,d); ax_cs.set_aspect('equal')
    ax_cs.axhline(0,color='#ffd700',lw=1.5,linestyle='--',label='Neutral Axis')
elif section_type == "I-Beam":
    bf,tf,hw,tw = dims['bf'],dims['tf'],dims['hw'],dims['tw']
    ax_cs.add_patch(mpatches.Rectangle((-bf/2,hw/2),bf,tf,color=mat_color+'44',ec='#00bfff',lw=2))
    ax_cs.add_patch(mpatches.Rectangle((-tw/2,-hw/2),tw,hw,color=mat_color+'44',ec='#00bfff',lw=2))
    ax_cs.add_patch(mpatches.Rectangle((-bf/2,-hw/2-tf),bf,tf,color=mat_color+'44',ec='#00bfff',lw=2))
    ax_cs.set_xlim(-bf,bf); ax_cs.set_ylim(-(hw+2*tf),(hw+2*tf))
    ax_cs.axhline(0,color='#ffd700',lw=1.5,linestyle='--',label='Neutral Axis')
elif section_type == "T-Beam":
    bf,tf,hw,tw = dims['bf'],dims['tf'],dims['hw'],dims['tw']
    yb = dims['y_bar']; ht = hw+tf
    ax_cs.add_patch(mpatches.Rectangle((-bf/2,hw-yb),bf,tf,color=mat_color+'44',ec='#00bfff',lw=2))
    ax_cs.add_patch(mpatches.Rectangle((-tw/2,-yb),tw,hw,color=mat_color+'44',ec='#00bfff',lw=2))
    ax_cs.set_xlim(-bf,bf); ax_cs.set_ylim(-ht,ht)
    ax_cs.axhline(0,color='#ffd700',lw=1.5,linestyle='--',label='Neutral Axis')
ax_cs.set_xlabel("Width (m)",color='#aaaaaa')
ax_cs.set_ylabel("Height (m)",color='#aaaaaa')
ax_cs.legend(facecolor='#0d2137',labelcolor='white',fontsize=7)
y_fib = np.linspace(-h_total/2,h_total/2,100)
sig_fib = (M_max*y_fib)/I/1e6
ax_sd.plot(sig_fib,y_fib,color='#ff6b35',lw=2.5)
ax_sd.fill_betweenx(y_fib,sig_fib,0,where=(sig_fib>0),color='#ff475733',label='Tension')
ax_sd.fill_betweenx(y_fib,sig_fib,0,where=(sig_fib<0),color='#00bfff33',label='Compression')
ax_sd.axvline(0,color='white',lw=0.8)
ax_sd.axhline(0,color='#ffd700',lw=1.5,linestyle='--',label='Neutral Axis')
ax_sd.set_xlabel("Stress (MPa)",color='#aaaaaa')
ax_sd.set_ylabel("Height (m)",color='#aaaaaa')
ax_sd.set_title("Stress Distribution",color='#00bfff')
ax_sd.legend(facecolor='#0d2137',labelcolor='white',fontsize=8)
plt.tight_layout()
st.pyplot(fig_cs)

# ════════════════════════════════════════════
# NEW FEATURE 3: MATERIAL COMPARISON
# ════════════════════════════════════════════
st.markdown("---")
st.markdown("### 🧪 Material Comparison Table")
mat_comparison = []
for mat_name, mat_props in material_data.items():
    E_m = mat_props["E"]
    y_m = mat_props["yield"]
    d_m = mat_props["density"]
    fos_m = (y_m/1e6) / sigma_max if sigma_max > 0 else 999
    defl_m = y_max * (E_val / E_m)  # scale deflection by E ratio
    weight_m = d_m * area * L * 9.81
    mat_comparison.append({
        "Material": mat_name,
        "E (GPa)": f"{E_m/1e9:.0f}",
        "Yield (MPa)": f"{y_m/1e6:.0f}",
        "FOS": f"{min(fos_m,999):.2f}",
        "Deflection (mm)": f"{defl_m:.3f}",
        "Weight (N)": f"{weight_m:.0f}",
        "Safe?": "✅" if fos_m > 2.0 else ("⚠️" if fos_m > 1.0 else "❌")
    })

import pandas as pd
df_mat = pd.DataFrame(mat_comparison)
st.dataframe(df_mat.set_index("Material"), width='stretch')

# ════════════════════════════════════════════
# NEW FEATURE 4: EXCEL EXPORT
# ════════════════════════════════════════════
st.markdown("---")
st.markdown("### 📊 Download Results")
col_d1, col_d2, col_d3 = st.columns(3)

# PNG Download
buf = io.BytesIO()
fig.savefig(buf,format='png',dpi=150,bbox_inches='tight',facecolor='#0a1628')
buf.seek(0)
col_d1.download_button("📥 Download Diagrams (PNG)", buf, "beam_analysis.png", "image/png")

# Excel Export
def generate_excel():
    wb = openpyxl.Workbook()
    
    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Summary"
    
    hdr_fill  = PatternFill("solid", fgColor="0D2137")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    val_fill  = PatternFill("solid", fgColor="1B3A5C")
    val_font  = Font(color="E0E0E0", size=10)
    safe_font = Font(bold=True, color="00C853", size=11)
    fail_font = Font(bold=True, color="FF4757", size=11)
    thin = Border(
        left=Side(style='thin', color='2A4A6C'),
        right=Side(style='thin', color='2A4A6C'),
        top=Side(style='thin', color='2A4A6C'),
        bottom=Side(style='thin', color='2A4A6C')
    )
    center = Alignment(horizontal='center', vertical='center')
    
    def set_cell(ws, row, col, value, fill=None, font=None, align=None, border=None):
        c = ws.cell(row=row, column=col, value=value)
        if fill:   c.fill = fill
        if font:   c.font = font
        if align:  c.alignment = align
        if border: c.border = border
        return c
    
    # Title
    ws1.merge_cells('A1:G1')
    set_cell(ws1, 1, 1, "2D Beam Stress & Deflection Analysis Report",
             PatternFill("solid", fgColor="001A33"),
             Font(bold=True, color="00BFFF", size=14), center)
    ws1.row_dimensions[1].height = 30
    
    # Input Parameters
    ws1.merge_cells('A3:G3')
    set_cell(ws1, 3, 1, "INPUT PARAMETERS", hdr_fill, hdr_font, center)
    
    params = [
        ("Beam Type", beam_type), ("Length (m)", L),
        ("Material", E_sel), ("Section Type", section_type),
        ("UDL (N/m)", w), ("E (GPa)", f"{E_val/1e9:.0f}"),
        ("Yield Strength (MPa)", f"{yield_MPa:.0f}"),
        ("Moment of Inertia (m⁴)", f"{I:.4e}"),
        ("Beam Weight (N)", f"{beam_weight:.1f}"),
    ]
    for i, (k, v) in enumerate(params, start=4):
        set_cell(ws1, i, 1, k, val_fill, Font(color="00D4AA", bold=True, size=10), border=thin)
        set_cell(ws1, i, 2, v, PatternFill("solid", fgColor="0A1628"), val_font, border=thin)
        ws1.column_dimensions['A'].width = 28
        ws1.column_dimensions['B'].width = 22
    
    row = 4 + len(params) + 1
    
    # Loads
    ws1.merge_cells(f'A{row}:G{row}')
    set_cell(ws1, row, 1, "APPLIED LOADS", hdr_fill, hdr_font, center)
    row += 1
    set_cell(ws1, row, 1, "Load #", hdr_fill, hdr_font, center, thin)
    set_cell(ws1, row, 2, "Force (N)", hdr_fill, hdr_font, center, thin)
    set_cell(ws1, row, 3, "Position (m)", hdr_fill, hdr_font, center, thin)
    row += 1
    for i, (P, a) in enumerate(loads, 1):
        set_cell(ws1, row, 1, f"P{i}", val_fill, val_font, center, thin)
        set_cell(ws1, row, 2, P, val_fill, val_font, center, thin)
        set_cell(ws1, row, 3, a, val_fill, val_font, center, thin)
        row += 1
    row += 1
    
    # Results
    ws1.merge_cells(f'A{row}:G{row}')
    set_cell(ws1, row, 1, "RESULTS", hdr_fill, hdr_font, center)
    row += 1
    
    results = [
        ("Reaction RA (N)", f"{RA:.2f}"),
        ("Reaction RB (N)", f"{RB:.2f}"),
        ("Max Bending Moment (N·m)", f"{M_max:.2f}"),
        ("Max Deflection (mm)", f"{y_max:.4f}"),
        ("Max Bending Stress (MPa)", f"{sigma_max:.2f}"),
        ("Yield Strength (MPa)", f"{yield_MPa:.0f}"),
        ("Factor of Safety", f"{FOS:.2f}"),
        ("Status", "SAFE ✅" if sigma_max < yield_MPa else "UNSAFE ❌"),
    ]
    for k, v in results:
        set_cell(ws1, row, 1, k, val_fill, Font(color="00D4AA", bold=True, size=10), border=thin)
        cell = set_cell(ws1, row, 2, v, PatternFill("solid", fgColor="0A1628"),
                        safe_font if "SAFE ✅" in str(v) else (fail_font if "UNSAFE" in str(v) else val_font),
                        border=thin)
        row += 1
    
    # ── Sheet 2: Raw Data ──
    ws2 = wb.create_sheet("SFD_BMD_Data")
    ws2.sheet_properties.tabColor = "00BFFF"
    
    headers = ["Position (m)", "Shear Force (N)", "Bending Moment (N·m)", "Deflection (mm)"]
    for col, h in enumerate(headers, 1):
        set_cell(ws2, 1, col, h, hdr_fill, hdr_font, center, thin)
        ws2.column_dimensions[chr(64+col)].width = 20
    
    step = max(1, len(x)//200)
    for row_i, idx in enumerate(range(0, len(x), step), start=2):
        vals = [round(float(x[idx]),4), round(float(V[idx]),4),
                round(float(M[idx]),4), round(float(y[idx]*1000),6)]
        for col, v in enumerate(vals, 1):
            fill = PatternFill("solid", fgColor="0A1628" if row_i%2==0 else "0D1F35")
            set_cell(ws2, row_i, col, v, fill, val_font, center, thin)
    
    # ── Sheet 3: Material Comparison ──
    ws3 = wb.create_sheet("Material_Comparison")
    ws3.sheet_properties.tabColor = "7C4DFF"
    
    mat_headers = ["Material", "E (GPa)", "Yield (MPa)", "Density (kg/m³)", "FOS", "Deflection (mm)", "Weight (N)", "Safe?"]
    for col, h in enumerate(mat_headers, 1):
        set_cell(ws3, 1, col, h, hdr_fill, hdr_font, center, thin)
        ws3.column_dimensions[chr(64+col)].width = 22
    
    for row_i, row_data in enumerate(mat_comparison, start=2):
        vals = [row_data["Material"], row_data["E (GPa)"], row_data["Yield (MPa)"],
                material_data[row_data["Material"]]["density"],
                row_data["FOS"], row_data["Deflection (mm)"], row_data["Weight (N)"], row_data["Safe?"]]
        for col, v in enumerate(vals, 1):
            fill = PatternFill("solid", fgColor="0A1628" if row_i%2==0 else "0D1F35")
            fnt = (safe_font if v == "✅" else fail_font if v == "❌" else val_font)
            set_cell(ws3, row_i, col, v, fill, fnt, center, thin)
    
    # Add Line chart for SFD
    chart = LineChart()
    chart.title = "Shear Force Diagram"
    chart.style = 10
    chart.y_axis.title = "Shear Force (N)"
    chart.x_axis.title = "Position"
    
    n_pts = min(50, len(x)//step)
    data_ref = Reference(ws2, min_col=2, min_row=1, max_row=n_pts+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.series[0].graphicalProperties.line.solidFill = "00BFFF"
    ws2.add_chart(chart, "F2")
    
    # BMD chart
    chart2 = LineChart()
    chart2.title = "Bending Moment Diagram"
    chart2.style = 10
    chart2.y_axis.title = "Moment (N·m)"
    data_ref2 = Reference(ws2, min_col=3, min_row=1, max_row=n_pts+1)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.series[0].graphicalProperties.line.solidFill = "FF6B35"
    ws2.add_chart(chart2, "F20")
    
    buf_xl = io.BytesIO()
    wb.save(buf_xl)
    buf_xl.seek(0)
    return buf_xl

if EXCEL_OK:
    if col_d2.button("📊 Generate Excel Report"):
        with st.spinner("Generating Excel..."):
            xl_buf = generate_excel()
        col_d2.download_button("📥 Download Excel", xl_buf,
                               "beam_analysis.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    col_d2.warning("Install openpyxl for Excel export")

# PDF Report
def generate_pdf(beam_type,L,loads,w,E,section_type,dims,I,c_dist,
                 RA,RB,M_max,y_max,sigma_max,yield_MPa,FOS,fig,warnings,suggestions,good):
    buf_pdf = io.BytesIO()
    doc = SimpleDocTemplate(buf_pdf,pagesize=A4,leftMargin=20*mm,rightMargin=20*mm,
                            topMargin=20*mm,bottomMargin=20*mm)
    title_style = ParagraphStyle('t',fontSize=18,textColor=colors.HexColor('#0D2137'),
                                  fontName='Helvetica-Bold',alignment=TA_CENTER,spaceAfter=6)
    sub_style   = ParagraphStyle('s',fontSize=11,textColor=colors.HexColor('#1B3A5C'),
                                  fontName='Helvetica',alignment=TA_CENTER,spaceAfter=12)
    head_style  = ParagraphStyle('h',fontSize=12,textColor=colors.HexColor('#0D2137'),
                                  fontName='Helvetica-Bold',spaceBefore=10,spaceAfter=4)
    body_style  = ParagraphStyle('b',fontSize=9,fontName='Helvetica',spaceAfter=3)
    story = []
    story.append(Paragraph("2D Beam Stress & Deflection Simulator Pro",title_style))
    story.append(Paragraph("Engineering Analysis Report",sub_style))
    story.append(Spacer(1,5*mm))
    story.append(Paragraph("1. Input Parameters",head_style))
    load_str = " | ".join([f"P{i+1}={p}N@{a}m" for i,(p,a) in enumerate(loads)])
    dim_str  = " | ".join([f"{k}={v}" for k,v in dims.items() if k!='y_bar'])
    params = [["Parameter","Value"],
              ["Beam Type",beam_type],["Length",f"{L}m"],
              ["Loads",load_str],["UDL",f"{w}N/m"],
              ["Material",E],["Section",section_type],
              ["Dimensions",dim_str],["I",f"{I:.4e} m4"],
              ["Beam Weight",f"{beam_weight:.1f} N"]]
    t = Table(params,colWidths=[70*mm,100*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#0D2137')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),9),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.HexColor('#F0F4FA'),colors.white]),
        ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#CBD5E0')),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
        ('LEFTPADDING',(0,0),(-1,-1),6)]))
    story.append(t); story.append(Spacer(1,5*mm))
    story.append(Paragraph("2. Results",head_style))
    safe_str = "SAFE" if sigma_max < yield_MPa else "UNSAFE"
    results = [["Result","Value","Unit"],
               ["RA",f"{RA:.2f}","N"],["RB",f"{RB:.2f}","N"],
               ["M_max",f"{M_max:.2f}","N.m"],["y_max",f"{y_max:.4f}","mm"],
               ["sigma_max",f"{sigma_max:.2f}","MPa"],["Yield",f"{yield_MPa:.0f}","MPa"],
               ["FOS",f"{FOS:.2f}","-"],["Status",safe_str,"-"]]
    t2 = Table(results,colWidths=[80*mm,60*mm,30*mm])
    sc = colors.HexColor('#00C853') if sigma_max < yield_MPa else colors.HexColor('#FF4757')
    t2.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#0D2137')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),9),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.HexColor('#F0F4FA'),colors.white]),
        ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#CBD5E0')),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
        ('LEFTPADDING',(0,0),(-1,-1),6),
        ('TEXTCOLOR',(1,-1),(1,-1),sc),('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold')]))
    story.append(t2); story.append(Spacer(1,5*mm))
    story.append(Paragraph("3. AI Engineering Suggestions",head_style))
    for warn in warnings: story.append(Paragraph(warn,body_style))
    for sugg in suggestions: story.append(Paragraph(sugg,body_style))
    for g in good: story.append(Paragraph(g,body_style))
    story.append(Spacer(1,5*mm))
    story.append(Paragraph("4. Engineering Diagrams",head_style))
    buf_img = io.BytesIO()
    fig.savefig(buf_img,format='png',dpi=120,bbox_inches='tight',facecolor='#0a1628')
    buf_img.seek(0)
    story.append(Image(buf_img,width=170*mm,height=120*mm))
    doc.build(story)
    buf_pdf.seek(0)
    return buf_pdf

if col_d3.button("📄 Generate PDF Report"):
    with st.spinner("Generating PDF..."):
        pdf_buf = generate_pdf(beam_type,L,loads,w,E_sel,section_type,dims,I,c_dist,
                               RA,RB,M_max,y_max,sigma_max,yield_MPa,FOS,fig,
                               warnings,suggestions,good)
    col_d3.download_button("📥 Download PDF Report",pdf_buf,"beam_report.pdf","application/pdf")

with st.expander("📚 Theory & Formulas"):
    st.markdown(f"""
    **{beam_type} | {section_type} Section | {E_sel}**
    - `I` = **{I:.4e} m⁴** | `c` = **{c_dist*1000:.1f} mm**
    - `RA` = **{RA:.1f} N** | `RB` = **{RB:.1f} N**
    - `M_max` = **{M_max:.1f} N·m** | `σ_max` = **{sigma_max:.2f} MPa**
    - `FOS` = **{FOS:.2f}** | `y_max` = **{y_max:.4f} mm**
    - `Beam Weight` = **{beam_weight:.1f} N** | `Density` = **{density} kg/m³**
    """)
